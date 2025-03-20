import openpyxl
from openpyxl.styles import Font
from io import BytesIO
from datetime import datetime
import streamlit as st
from builtins import int  # Explicitly import int

# Streamlit UI setup
st.title('Filter Excel Data')
st.write('Upload an Excel file to include all data and ensure RO Remarks are present.')

# File uploader widget
uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

# Function to process the Excel file without using pandas
def process_excel_file(file):
    # Load the Excel file
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    # Read the headers from the first row (assumes first row is headers)
    headers = [cell.value for cell in sheet[1]]
    
    # Define the exact header order you want
    columns_order = ['Date Submitted', 'CYCLE', 'Customer Number', "Cardholder's Name", 'OB/AOD', 'Source of Income', 'Reason to Avail', 
                     'Email', 'Date sent to RO', 'RO Remarks', 'SENT TO CLIENT', 'SIGNED SBC FORWARDED TO RO', 'CH STATUS', 'FINONE ID']

    # Check if all required columns are in the sheet
    missing_cols = [col for col in columns_order if col not in headers]
    if missing_cols:
        st.write(f"Warning: Missing columns - {', '.join(missing_cols)}")
        return

    # Reorder the columns
    column_indices = [headers.index(col) for col in columns_order]
    
    # Process data rows
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        new_row = [row[i] for i in column_indices]
        
        # Customer Number: Add leading zero and remove .0
        if isinstance(new_row[2], (int, float)):
            new_row[2] = f"0{int(new_row[2])}" if new_row[2] else '0000000000'
        
        # Fill blanks in 'Date sent to RO' with the current date
        if not new_row[8]:  # Assuming 'Date sent to RO' is at index 8
            new_row[8] = datetime.today().strftime('%Y/%m/%d')

        # Rename 'Date Submitted' to 'Call Date' and format it
        new_row[0] = new_row[0].strftime('%m/%d/%Y') if isinstance(new_row[0], datetime) else new_row[0]
        
        # Add PTP Amount (default value 0)
        new_row.insert(5, 0)  # Add 'PTP Amount' after 'OB/AOD'
        
        # Append the processed row
        data.append(new_row)

    return data

# Process the uploaded file
if uploaded_file:
    # Get processed data
    data = process_excel_file(uploaded_file)
    
    if data:
        # Display the filtered data
        st.write("Filtered Data (Including RO Remarks, Call Date, and PTP Amount):")
        st.dataframe(data)

        # Group the data by 'CYCLE' (column index 1)
        grouped_data = {}
        for row in data:
            cycle = row[1]
            if cycle not in grouped_data:
                grouped_data[cycle] = []
            grouped_data[cycle].append(row)

        # Function to create an Excel file with autofitted columns and bold headers
        def create_excel_with_formatting(group):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Filtered Data"

            # Add the header row
            headers = ['Call Date', 'CYCLE', 'Customer Number', "Cardholder's Name", 'OB/AOD', 'PTP Amount', 'Source of Income', 'Reason to Avail',
                       'Email', 'Date sent to RO', 'RO Remarks', 'SENT TO CLIENT', 'SIGNED SBC FORWARDED TO RO', 'CH STATUS']
            ws.append(headers)
            
            # Add data rows
            for row in group:
                ws.append(row)

            # Make header bold
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Autofit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Save to a BytesIO object for download
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        # Allow user to download data for each cycle
        current_filename_date = datetime.today().strftime('%m%d')
        
        for cycle, group in grouped_data.items():
            excel_data = create_excel_with_formatting(group)
            file_name = f"{cycle}*Conso list_SBC Request_Results ({current_filename_date}).xlsx"
            
            st.download_button(
                label=f"Download Cycle {cycle} Data (Excel)",
                data=excel_data,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.write("No data found to process.")
